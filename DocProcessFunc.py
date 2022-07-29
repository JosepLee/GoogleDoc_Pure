import sys

from tkinter.tix import COLUMN
from tokenize import cookie_re
import shutil
import copy



import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

import json
import os
from win32com.client import Dispatch
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment,PatternFill
from colorama import init
init(autoreset=True)

# color for 5 dimension of quizs
quiz_colors = ["f28e86", "8cb5f9", "7ad694", "8ed7dd", "fdd868"]
# colors for members
member_colors = ["1e79b2", "fe7f0c", "2d9f30", "d42826", "9368ba", "ff0000"]
def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(os.path.abspath(filename))
    xlBook.Save()
    xlBook.Close()

'''
json_process()
处理json文件，读取并形成名单和名字权重对应
'''
def json_Process(file_Path):
    # Json process func:
    # 如果没有RawData文件夹，报错，没有提供原始数据，原始数据应该放在RawData文件夹内
    if not os.path.exists('RawData'):
        print(
            "\033[0;31;40mError,you don't have the RawData folder. I make one for you,please fill the rawdata into the RawData folder.\033[0m")
        os.mkdir('RawData')
        input("Press any key to stop processing.")
        sys.exit()
    print('Processing the json file: {}.'.format(file_Path))
    try:
        with open(file_Path) as f:
            data = json.load(f)
    except FileNotFoundError:
        print("\033[0;31;40mError: You have no {}.\033[0m".format(file_Path))
        input("Press any key to stop processing.")
        sys.exit()

    member = data['member']
    weight = data['weight']

    name_Job_Set = {}
    job_Weight_set = {}
    name_list = []
    weight_Check = 0
    error = 0
    weight_Check_List = []
    jobList = []
    for it in weight:
        jobList.append(it['job'])
    for it in member:
        if it['job'] not in jobList:
            print(
                "\033[0;31;40mError: {}\'s job:{} is not exist in Job list,please check your json file.\033[0m".format(
                    it['name'], it['job']))
            error = 1
    for it in member:
        name_Job_Set[it['name']] = it['job']
        name_list.append(it['name'])
    for it in weight:
        job_Weight_set[it['job']] = it['weight']
        if len(it['weight']) != 5:
            print(
                '\033[0;31;40mWeight Error:Weight of job:{} list length is not 5,please check your json file.\033[0m'.format(
                    it['job']))
            error = 1
        for its in it['weight']:
            weight_Check = weight_Check + its
        if int(weight_Check) - 1 > 0.01:
            print(
                '\033[0;31;40mWeight Error: The weight sum of job:{} is not 1,please check your json file.\033[0m'.format(
                    it['job']))
            error = 1
        weight_Check = 0

    # open the workbook
    #TODO 直接处理出一个name weight set
    weight_name={}
    for name in name_list:
        weight_name[name]=job_Weight_set[name_Job_Set[name]]
    print('Json Processing finished.')
    if error:
        return [], [], []
    else:
        return name_list, weight_name

'''
把rawdata文件夹中的文件列表读取处理，并且处理是否有不同月份报表人员变化的情况
'''
def RawDataInit(workBookPaths):

    jsonList=[]
    tableList=[]
    findxlsx = 0
    findjson = 0
    for it in workBookPaths:
        if it.find('xlsx') != -1:
            findxlsx = 1
            tableList.append(it)

        if it.find('json') != -1:
            findjson = 1
            jsonList.append(it)

    if (findxlsx == 0):
        print('\033[0;31;40mError:There is no table in RawData.\033[0m')
        input("“Press Any Key to close terminal.”")
        sys.exit()

        # 此处加上异常处理!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    if (findjson == 0):
        print('\033[0;31;40mError:There is no json in RawData.\033[0m')
        input("“Press Any Key to close terminal.”")
        sys.exit()
    elif (len(workBookPaths) == 1):
        print('Start to process tables.')
        print("Table name is: {}".format(workBookPaths[:]))

    else:
        print('Start to process tables.'.format(len(workBookPaths)))
        print("Files in RawData: {}".format(workBookPaths[:]))

    namedict={}
    weightdict={}
    Pmode=''
    while Pmode!='y' and Pmode!='n':
        Pmode = input('Do you have changes on team member,job or weight in different month?[y/n]')
    if Pmode=='y':
        print('You have selected team change mode, make sure you have json for every month\'s table.')
        print('For example,you should have MR2201.xlsx and MR2201.json. And you can not have any other json')

        for it in tableList:
            tablename=it[it.find('MR'):it.find('MR') + 6]
            namedict[tablename],weightdict[tablename]=json_Process('RawData\\{}.json'.format(tablename))
    else:
        for it in tableList:
            tablename = it[it.find('MR'):it.find('MR') + 6]
            namedict[tablename],weightdict[tablename]=json_Process('RawData\\team.json')
    tableList.reverse()
    return tableList,namedict,weightdict

#TODO make rateClass private
rateClass = ['Name', 'Output', "Productivity", "Self-driven", "Cooperation", "Innovation"]

'''
用来画每月个人报表 ifsum控制是不是画总表
'''
def chartDrawer(tableName, sheetNum, name, skiprows, ifsum, ifweight,nameList):
    if ifsum == 1:
        rowNum = len(nameList) + 1
    else:
        rowNum = 2
    workbookprocessed = pd.read_excel(tableName, sheet_name=sheetNum, skiprows=skiprows, index_col=0,
                                      usecols=[0, 1, 2, 3, 4, 5])

    # print(workbookprocessed)
    # print(list(workbookprocessed.index))
    # print(list(workbookprocessed.columns))
    a = tableName.find('MR')
    tableSubName = tableName[a:a + 6]

    x_row = np.arange(5)
    x_label = workbookprocessed.columns.to_numpy()

    y_col = []
    for i in range(rowNum):
        a = workbookprocessed.iloc[i].to_numpy()
        a = np.around(a, 1)
        y_col.append(a)

    # print(y_col[1])

    rect1 = [0.1, 0.5, 0.8, 0.45]
    rect2 = [0.25, 0.0, 0.7, 0.5]
    fig = plt.figure(figsize=(10, 8))

    ax1 = plt.axes(rect1)
    ax2 = plt.axes(rect2)

    ax1.set_xticks(x_row)
    ax1.set_yticks([4, 8])
    ax1.set_xticklabels(x_label, rotation=0, fontsize=15)

    # ax1.set_xlabel("issue",fontsize=15)
    ax1.set_ylabel("score", fontsize=15)

    ax1.grid(True, linestyle=':')

    bar_width = 0.125
    C_colors = ["#f28e86", "#8cb5f9", "#7ad694", "#8ed7dd", "#fdd868", ]
    R_colors = ["#1e79b2", "#fe7f0c", "#2d9f30", "#d42826", "#9368ba", "#1e79b2", "#fe7f0c", "#2d9f30", "#d42826",
                "#9368ba", "#ff00ff"]
    ls = []
    if ifsum:
        for i in range(rowNum - 1):
            l1 = ax1.bar(x_row - 2 * bar_width + bar_width * i, y_col[i], bar_width, color=R_colors[i])
            ax1.bar_label(l1, fmt='%.1f', label_type='edge')
            ls.append(l1)
        l2 = ax1.plot(x_row, y_col[-1], 'o-', color="#54164a")
        for x, y in zip(x_row, y_col[-1]):
            ax1.text(x + 0.10, y - 0.6, '%.1f' % y, ha='center', va='bottom', fontsize=15, color="#54164a")
        ax1.set_title('Sum Chart', fontsize=12)
    else:
        for i in range(rowNum - 1):
            l1 = ax1.bar(x_row, y_col[i], bar_width, color=R_colors[i])
            ax1.bar_label(l1, fmt='%.1f', label_type='edge')
            ls.append(l1)
        l2 = ax1.plot(x_row, y_col[-1], 'o-', color="#54164a")
        for x, y in zip(x_row, y_col[-1]):
            ax1.text(x + 0.10, y - 0.6, '%.1f' % y, ha='center', va='bottom', fontsize=15, color="#54164a")
        ax1.set_title(name, fontsize=12)
    labels = []
    labels = nameList.copy()
    ax1.legend(handles=ls, labels=labels, loc=(1, 0.5))
    cellTexts = [workbookprocessed]

    colLabel = list(workbookprocessed.columns)
    rowLabel = list(workbookprocessed.index)

    cellTexts = workbookprocessed.to_numpy()
    cellTexts = np.around(cellTexts, 1)
    # print(cellTexts)
    # print(workbookprocessed.columns)
    # print(workbookprocessed.index)
    tab = ax2.table(cellText=cellTexts,  # 简单理解为表示表格里的数据
                    colWidths=[0.2] * 5,  # 每个小格子的宽度 * 个数，要对应相应个数
                    colLabels=colLabel,  # 每列的名称
                    colColours=C_colors,
                    rowColours=R_colors,  # 每列名称颜色
                    rowLabels=rowLabel,  # 每行的名称（从列名称的下一行开始）
                    rowLoc="center",
                    cellLoc="center",  # 行名称的对齐方式
                    loc="center"  # 表格所在位置
                    )
    tab.auto_set_font_size(False)
    tab.set_fontsize(15)
    tab.scale(1, 2)
    ax2.axis('off')

    plt.close()

    if ifweight == 0:
        if not os.path.exists("cache\\Chart\\"):
            os.mkdir("cache/Chart")
        fig.savefig('cache\\Chart\\{}_{}.jpg'.format(tableSubName, sheetNum),bbox_inches='tight')
    else:
        if not os.path.exists("cache\\weighted\\Chart\\"):
            os.mkdir("cache/weighted/Chart")
        fig.savefig('cache\\weighted\\Chart\\{}_{}.jpg'.format(tableSubName, sheetNum),bbox_inches='tight')

'''
用来画个人对比表
'''
def chartDrawerSum(tableName, sheetNum, name, skiprows, nrow, historynum, ifweight,nameList):
    rowNum = historynum
    workbookprocessed = pd.read_excel(tableName, sheet_name=sheetNum, skiprows=skiprows, nrows=nrow, index_col=0,
                                      usecols=[0, 1, 2, 3, 4, 5])

    # print(workbookprocessed)
    # print(list(workbookprocessed.index))
    # print(list(workbookprocessed.columns))

    x_row = np.arange(5)
    x_label = workbookprocessed.columns.to_numpy()

    y_col = []
    for i in range(rowNum):
        a = workbookprocessed.iloc[i].to_numpy()
        a = np.around(a, 1)
        y_col.append(a)
    y_col.reverse()
    # print(y_col[1])

    rect1 = [0.1, 0.5, 0.8, 0.45]
    rect2 = [0.25, 0.0, 0.7, 0.5]

    fig = plt.figure(figsize=(10, 8))
    ax1 = plt.axes(rect1)
    ax2 = plt.axes(rect2)

    ax1.set_xticks(x_row)
    ax1.set_yticks([4, 8])
    ax1.set_xticklabels(x_label, rotation=0, fontsize=15)

    # ax1.set_xlabel("issue",fontsize=15)
    ax1.set_ylabel("score", fontsize=15)

    ax1.grid(True, linestyle=':')

    bar_width = 0.125
    C_colors = ["#f28e86", "#8cb5f9", "#7ad694", "#8ed7dd", "#fdd868"]
    R_colors = ["#1e79b2", "#fe7f0c", "#2d9f30", "#d42826", "#9368ba", "#1e79b2", "#fe7f0c", "#2d9f30", "#d42826",
                "#9368ba", "#ff00ff"]
    ls = []

    for i in range(rowNum):
        l1 = ax1.bar(x_row - 2 * bar_width + bar_width * i, y_col[i], bar_width, color=R_colors[i])
        ax1.bar_label(l1, fmt='%.1f', label_type='edge')
        ls.append(l1)
    ax1.set_title(name, fontsize=15)

    labels = []
    labels = nameList.copy()
    idx = list(workbookprocessed.index)[:]
    chartlegend = []
    for it in idx:
        chartlegend.append(it)
    chartlegend.reverse()
    ax1.legend(handles=ls, labels=chartlegend, loc=(1, 0.5))

    cellTexts = [workbookprocessed]

    colLabel = list(workbookprocessed.columns)
    rowLabel = list(workbookprocessed.index)
    rowLabel.reverse()
    cellTexts = workbookprocessed.to_numpy()
    cellTexts = np.flip(cellTexts, 0)
    cellTexts = np.around(cellTexts, 1)
    # print(cellTexts)
    # print(workbookprocessed)

    tab = ax2.table(cellText=cellTexts,  # 简单理解为表示表格里的数据
                    colWidths=[0.2] * 5,  # 每个小格子的宽度 * 个数，要对应相应个数
                    colLabels=colLabel,  # 每列的名称
                    colColours=C_colors,
                    rowColours=R_colors,  # 每列名称颜色
                    rowLabels=rowLabel,  # 每行的名称（从列名称的下一行开始）
                    rowLoc="center",
                    cellLoc="center",  # 行名称的对齐方式
                    loc="center"  # 表格所在位置
                    )
    tab.auto_set_font_size(False)
    tab.set_fontsize(15)
    tab.scale(1, 2)
    ax2.axis('off')

    plt.close()

    if ifweight == 0:
        if not os.path.exists("cache\\ChartHistory\\"):
            os.mkdir("cache/ChartHistory")
        fig.savefig('cache\\ChartHistory\\history_{}.jpg'.format(sheetNum),bbox_inches='tight')
    else:
        if not os.path.exists("cache\\weighted\\ChartHistory\\"):
            os.mkdir("cache/weighted/ChartHistory")
        fig.savefig('cache\\weighted\\ChartHistory\\history_{}.jpg'.format(sheetNum),bbox_inches='tight')

'''
处理rawdata表格，然后生成cache信息
'''
def MainChartProcess(workbook, workbookname,ifweight,nameListsum,weightListsum):
    # 新建Cache文件夹判断
    if ifweight=='y':
        if not os.path.exists('cache\\weighted\\'):
            os.mkdir('cache/weighted')
        if not os.path.exists('cache\\weighted\\processedTable'):
            os.mkdir('cache/weighted/processedTable')
        if not os.path.exists('cache\\weighted\\Chart'):
            os.mkdir('cache/weighted/Chart')
        if not os.path.exists('cache\\weighted\\ChartHistory'):
            os.mkdir('cache/weighted/ChartHistory')
        tableCacheName = 'cache\\weighted\\processedTable\\{}_final.xlsx'.format(workbookname)
    else:
        if not os.path.exists('cache'):
            os.mkdir('cache')
        if not os.path.exists('cache\\processedTable'):
            os.mkdir('cache/processedTable')
        if not os.path.exists('cache\\Chart'):
            os.mkdir('cache/Chart')
        if not os.path.exists('cache\\ChartHistory'):
            os.mkdir('cache/ChartHistory')
        tableCacheName = 'cache\\processedTable\\{}_final.xlsx'.format(workbookname)

    # 预处理,将提交的表单重新排序和json中的姓名顺序相同,方便后续去除自我评价
    arrangeWorksheet = workbook.create_sheet(title='Preprocess')


    #TODO 解耦合nameList
    nameList=nameListsum[workbookname]
    weightListsums=weightListsum[workbookname]
    memberNum = len(nameList)
    #TODO 把yangzi的匿名给解耦合
    Anonymous = 0
    # 判断是否匿名
    if workbook.worksheets[0].cell(row=1, column=2).value.find('You') == -1:
        Anonymous = 1
    # 如果是匿名的，则直接copy一份原始数据到processed
    if Anonymous == 1:
        for row in workbook.worksheets[0].iter_rows():
            arrangeWorksheet.append(cell.value for cell in row)
    if Anonymous == 0:
        arrangeWorksheet.append(cell.value for cell in workbook.worksheets[0][1])
        for name in nameList:
            calc = 0
            stop = 0
            find = 0
            for row in workbook.worksheets[0].iter_rows():
                if row[1].value == None:
                    stop = 1
                    break
                if name in row[1].value:
                    arrangeWorksheet.append(cell.value for cell in row)
                    find = 1
                calc = calc + 1
            if not find:
                print(
                    '\033[0;31;40mError:Member {} is not found in table,check if name in Json file is different from the google drive or wrong input in Json file.\033[0m'.format(
                        name))
                print('process Terminated')
                return -1, -1, -1, -1, -1, -1
            if stop:
                continue

        if calc != len(nameList) + 1:
            if calc < len(nameList) + 1:
                print(
                    '\033[0;31;40mError:Json namelist is larger than the amount of info in table,maybe someone didn\'t submit the google drive,or you reinput some member.\033[0m')
                print('process Terminated')
                return -1, -1, -1, -1, -1, -1
            if calc > len(nameList) + 1:
                print(
                    '\033[0;31;40mError:Json namelist is smaller than the amount of info in table,maybe you didn\'t fill all member in Json file.\033[0m')
                print('process Terminated')
                return -1, -1, -1, -1, -1, -1

    # 主处理模块,将preprocess表单转化成每个人的分表单,并清晰显示,外加后续处理.
    # 横行标为问题,列标为评价人,和namelist顺序相同
    dimensionQuizNum = []
    count = 0
    go = 0

    for name in nameList:
        quizMark = 1  # quiz内容的起始行

        if Anonymous == 0:
            commentMark = int((arrangeWorksheet.max_column - 2) / len(nameList)) - 4
        if Anonymous == 1:
            commentMark = int((arrangeWorksheet.max_column - 1) / len(nameList)) - 4
        commentMarkReturn = commentMark  # 评论内容的起始行
        newWorksheet = workbook.create_sheet(title=name)

        for col in arrangeWorksheet.iter_cols():
            if name in col[0].value and col[0].value[-1] == ']':
                quiz = col[0].value
                quiz = quiz[:quiz.find('[')]
                colz = []
                colz.append(quiz)
                for cell in col[1:]:
                    colz.append(cell.value)
                for i in range(len(colz)):
                    newWorksheet.cell(quizMark, i + 1, colz[i])
                quizMark = quizMark + 1
                count = count + 1
                flag = 0
            if name in col[0].value and col[0].value[-1] != ']':
                quiz = col[0].value
                colz = []
                colz.append(quiz)
                for cell in col[1:]:
                    colz.append(cell.value)
                for i in range(len(colz)):
                    newWorksheet.cell(commentMark, i + 1, colz[i])
                commentMark = commentMark + 1
                if go == 0:
                    dimensionQuizNum.append(count)
                    count = 0
            if go == 0:
                commentMarkEnd = commentMark

        go = go + 1

    quizend = [0, 0, 0, 0, 0]
    quizstart = [1, 1, 1, 1, 1]
    for i in range(1, len(dimensionQuizNum) + 1):
        for j in range(i):
            quizend[i - 1] = quizend[i - 1] + dimensionQuizNum[j]
    for i in range(len(dimensionQuizNum)):
        for j in range(i):
            quizstart[i] = quizstart[i] + dimensionQuizNum[j]
    totalMark = memberNum + 6
    mainChartMarkReturn = totalMark
    # sheetMark = commentMarkEnd+5

    # 主表表头写入
    for i in range(len(rateClass)):
        workbook.worksheets[0].cell(totalMark, i + 1, rateClass[i])
    totalMark = totalMark + 1

    # 删除自我评价评分行
    for i, name in enumerate(nameList):

        weightList=weightListsums[name]
        # 进行美化
        workbook[name].column_dimensions['A'].width = 100
        for j in range(1, memberNum):
            workbook[name].column_dimensions[chr(ord('A') + j)].width = 30
            # print(chr(ord('A')+i))
        # workbook[name].column_dimensions['C'].width = 30
        # workbook[name].column_dimensions['D'].width = 30
        # workbook[name].column_dimensions['E'].width = 30
        # workbook[name].column_dimensions['F'].width = 30
        # 删除对自己的评价
        if Anonymous == 0:
            workbook[name].delete_cols(int(i + 2))
        # 写评分类题头
        avgMarkReturn = commentMarkEnd + 5
        for i in range(len(rateClass)):
            workbook[name].cell(avgMarkReturn - 1, i + 1, rateClass[i])
        # 求出平均值,填写到总表处。

        # 计算每个分chart的问题数量
        if Anonymous == 0:
            avgReigonIndex = chr(ord("B") + memberNum - 2)
        if Anonymous == 1:
            avgReigonIndex = chr(ord("B") + memberNum - 1)

        #如果带权，就带权重计算均值
        if ifweight=='y':
            workbook[name].cell(row=avgMarkReturn, column=1).value = name
            workbook[name].cell(row=avgMarkReturn,
                                column=2).value = f"=AVERAGE(B{quizstart[0]}:{avgReigonIndex}{quizend[0]})*{weightList[0]}"
            workbook[name].cell(row=avgMarkReturn,
                                column=3).value = f"=AVERAGE(B{quizstart[1]}:{avgReigonIndex}{quizend[1]})*{weightList[1]}"
            workbook[name].cell(row=avgMarkReturn,
                                column=4).value = f"=AVERAGE(B{quizstart[2]}:{avgReigonIndex}{quizend[2]})*{weightList[2]}"
            workbook[name].cell(row=avgMarkReturn,
                                column=5).value = f"=AVERAGE(B{quizstart[3]}:{avgReigonIndex}{quizend[3]})*{weightList[3]}"
            workbook[name].cell(row=avgMarkReturn,
                                column=6).value = f"=AVERAGE(B{quizstart[4]}:{avgReigonIndex}{quizend[4]})*{weightList[4]}"
        else:
            workbook[name].cell(row=avgMarkReturn, column=1).value = name
            workbook[name].cell(row=avgMarkReturn,
                                column=2).value = f"=AVERAGE(B{quizstart[0]}:{avgReigonIndex}{quizend[0]})"
            workbook[name].cell(row=avgMarkReturn,
                                column=3).value = f"=AVERAGE(B{quizstart[1]}:{avgReigonIndex}{quizend[1]})"
            workbook[name].cell(row=avgMarkReturn,
                                column=4).value = f"=AVERAGE(B{quizstart[2]}:{avgReigonIndex}{quizend[2]})"
            workbook[name].cell(row=avgMarkReturn,
                                column=5).value = f"=AVERAGE(B{quizstart[3]}:{avgReigonIndex}{quizend[3]})"
            workbook[name].cell(row=avgMarkReturn,
                                column=6).value = f"=AVERAGE(B{quizstart[4]}:{avgReigonIndex}{quizend[4]})"

    # 处理Openpyxl导致的公式求平均无法更新值到表的问题
    workbook.save(tableCacheName)
    just_open(tableCacheName)

    wb = openpyxl.load_workbook(tableCacheName, data_only=True)

    # 把每个分表的平均值写到总表，同时计算各项得分平均值，再反馈平均值给各个分表
    avgList = [0, 0, 0, 0, 0]
    for i, name in enumerate(nameList):
        # 各成员分值写入总表，并计算平均值
        workbook.worksheets[0].cell(row=totalMark, column=1).value = name
        for i in range(5):
            workbook.worksheets[0].cell(row=totalMark, column=2 + i).value = wb[name].cell(row=avgMarkReturn,
                                                                                           column=i + 2).value
        for i in range(5):
            try:
                # print(wb[name].cell(row=avgMarkReturn, column=i+2).value)
                avgList[i] = avgList[i] + wb[name].cell(row=avgMarkReturn, column=i + 2).value
            except TypeError:
                print(
                    "\033[0;31;40mError: You have wrong name in online sheet. Example:Li Zhi Xing was typed as Lizx. Please check the google sheet.\033[0m")
                input('Press any key to stop process.')
                sys.exit()

        totalMark = totalMark + 1

    # 把平均值写入总表
    for i, num in enumerate(avgList):
        avgList[i] = num / len(nameList)
    workbook.worksheets[0].cell(row=totalMark, column=1).value = "Average"
    for i in range(2, 7):
        workbook.worksheets[0].cell(row=totalMark, column=i).value = avgList[i - 2]

    # 把平均值写入各分表
    for name in nameList:
        workbook[name].cell(row=avgMarkReturn + 1, column=1).value = "Average"
        for i in range(2, 7):
            workbook[name].cell(row=avgMarkReturn + 1, column=i).value = avgList[i - 2]

    workbook.save(tableCacheName)
    just_open(tableCacheName)

    if ifweight=='y':
        # 画图主表图部分
        chartDrawer(tableCacheName, 0, '', mainChartMarkReturn - 1, 1, 1,nameList)
        # 画图分表图部分
        for sheetNum in range(2, memberNum + 2):
            chartDrawer(tableCacheName, sheetNum, nameList[sheetNum - 2], avgMarkReturn - 2, 0, 1,nameList)
    else:
        # 画图主表图部分
        chartDrawer(tableCacheName, 0, '', mainChartMarkReturn - 1, 1, 0, nameList)
        # 画图分表图部分
        for sheetNum in range(2, memberNum + 2):
            chartDrawer(tableCacheName, sheetNum, nameList[sheetNum - 2], avgMarkReturn - 2, 0, 0, nameList)


    return dimensionQuizNum, quizstart, quizend, mainChartMarkReturn, commentMarkReturn, avgMarkReturn - 1, Anonymous
    # 分表绘图，函数循环调用即可

'''
把cache信息汇总成总表
'''
def sumTableGen(quizNumberL, quizmarkstartL, returnAvgL, returnCommentL, Anonymous,ifweight,nameListsum):
    resultdir = 'result\\'
    if not os.path.exists(resultdir):
        os.mkdir(resultdir)
    if ifweight=='y':
        if not os.path.exists('result\\PersonalReportWeighted'):
            os.mkdir('result/PersonalReportWeighted')
        processedWorkBookPaths = os.listdir('cache\\weighted\\processedTable\\')
        processedWorkBookPaths.reverse()
    else:
        if not os.path.exists('result\\PersonalReport'):
            os.mkdir('result/PersonalReport')
        processedWorkBookPaths = os.listdir('cache\\processedTable\\')
        processedWorkBookPaths.reverse()
    # 先新建一个表，作为结果经理总表
    sumChart = openpyxl.Workbook()
    # 然后建好总sheet和各分sheet
    sumSheet = sumChart.active
    sumSheet.title = 'SumSheet'

    #TODO 处理总表情况
    lenName=0
    nameListSuper=[]
    for it in nameListsum:
        if len(nameListsum[it])>lenName:
            nameListSuper=nameListsum[it]
            lenName=len(nameListsum[it])


    for name in nameListSuper:
        ws = sumChart.create_sheet(name)

    # 读取处理后的表格文件夹
    if ifweight=='y':
        saveTableName='sumTable_weighted.xlsx'
        sumChart.save(resultdir + saveTableName)
        MainWorkBook = openpyxl.load_workbook(resultdir + 'sumTable_weighted.xlsx', data_only=True)
        cacheDir='cache\\weighted\\'
    else:
        saveTableName='sumTable.xlsx'
        sumChart.save(resultdir + saveTableName)
        MainWorkBook = openpyxl.load_workbook(resultdir + 'sumTable.xlsx', data_only=True)
        cacheDir= 'cache\\'

    print('Sum table process finished.')

    sumChartRowMark = 1
    sheetRowMarkLoop = 50 + len(processedWorkBookPaths)

    sheetRowMark_List = []

    LoopTime = 0


    # 汇总形成总表
    for marknum, it in enumerate(processedWorkBookPaths):

        #TODO 此处有it，得动态输入列表
        # 赋值动态mark
        itName=it[:6]
        nameList=nameListsum[itName]
        quizNumber = quizNumberL[marknum]
        quizmarkstart = quizmarkstartL[marknum]
        returnAvg = returnAvgL[marknum]
        returnComment = returnCommentL[marknum]
        # 打开cache表
        processingWorkBook = openpyxl.load_workbook(cacheDir+'processedTable\\' + it, data_only=True)
        # 复制信息到总sheet表

        # 合并单元格做标题
        MainWorkBook.worksheets[0].merge_cells("A{}:L{}".format(sumChartRowMark, sumChartRowMark))
        # 调整标题格样式
        MainWorkBook.worksheets[0].cell(row=sumChartRowMark, column=1).value = it[:it.find('_')]
        MainWorkBook.worksheets[0].cell(row=sumChartRowMark, column=1).font = Font(size=30, bold=True)
        MainWorkBook.worksheets[0].cell(row=sumChartRowMark, column=1).alignment = Alignment(horizontal="center",
                                                                                             vertical="center",
                                                                                             wrap_text=True)
        MainWorkBook.worksheets[0].row_dimensions[sumChartRowMark].height = 30
        # 加图,2-一定大小的行，这个要动态调整？
        picPath = cacheDir+'Chart\\' + it[:it.find('_')] + '_' + '0' + '.jpg'
        img = Image(picPath)  # 选择你的图片
        MainWorkBook.worksheets[0].add_image(img, 'A{}'.format(sumChartRowMark + 2))

        MainWorkBook.save(resultdir + saveTableName)
        # 加图,2-一定大小的行，这个要动态调整？
        sumChartRowMark = sumChartRowMark + 46

        # 复制源信息
        loopTime = 0

        for row in processingWorkBook.worksheets[0].iter_rows():
            for i in range(1, processingWorkBook.worksheets[0].max_column + 1):
                MainWorkBook.worksheets[0].cell(row=sumChartRowMark, column=i).value = row[i - 1].value
            sumChartRowMark = sumChartRowMark + 1
            loopTime = loopTime + 1
            if loopTime == 3 + len(nameList):
                break

        MainWorkBook.save(resultdir + saveTableName)
        sumChartRowMark = sumChartRowMark + 3
        sheetRowMark_List.append(sumChartRowMark)

        # 分表处理

        for i, name in enumerate(nameList):
            sheetRowMark = sheetRowMarkLoop
            presentSheet = MainWorkBook.worksheets[i + 1]
            presentOriginal = processingWorkBook.worksheets[i + 2]
            for l in range(1, len(nameList) + 1):
                presentSheet.column_dimensions[chr(ord('A') + l)].width = 30
            # 合并单元格做标题
            presentSheet = MainWorkBook.worksheets[i + 1]
            presentOriginal = processingWorkBook.worksheets[i + 2]
            presentSheet.merge_cells("A{}:F{}".format(sheetRowMark, sheetRowMark))
            # 调整标题格样式
            presentSheet.cell(row=sheetRowMark, column=1).value = it[:it.find('_')]
            presentSheet.cell(row=sheetRowMark, column=1).font = Font(size=30, bold=True)
            presentSheet.cell(row=sheetRowMark, column=1).alignment = Alignment(horizontal="center",
                                                                                vertical="center",
                                                                                wrap_text=True)
            presentSheet.row_dimensions[sheetRowMark].height = 30
            # 加图,2-一定大小的行，这个要动态调整？
            picPath = cacheDir+'Chart\\' + it[:it.find('_')] + '_' + '{}'.format(i + 2) + '.jpg'
            img = Image(picPath)  # 选择你的图片
            presentSheet.add_image(img, 'A{}'.format(sheetRowMark + 2))

            # MainWorkBook.save(resultdir + 'sumTable.xlsx')

            # 加图,2-一定大小的行，这个要动态调整？
            sheetRowMark = sheetRowMark + 46
            # 对应标题的对应图：ex MR2207

            font = Font(size=24, bold=True)
            alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
            # pattern_fill = PatternFill(fill_type="solid", fgColor="f28e86")
            # 调整表格长宽
            presentSheet.column_dimensions['A'].width = 100

            lengthcomment = 0
            if Anonymous:
                lengthcomment = len(nameList) + 2
            else:
                lengthcomment = len(nameList) + 1
            # 复制评论
            for k in range(5):  # row
                pattern_fill = PatternFill(fill_type="solid", fgColor=member_colors[k])
                presentSheet.row_dimensions[sheetRowMark].height = 175

                for j in range(1, lengthcomment):  # col
                    presentSheet.cell(row=sheetRowMark, column=j).value = presentOriginal.cell(row=returnComment + k,
                                                                                               column=j).value
                    presentSheet.cell(row=sheetRowMark, column=j).font = font
                    if (j != 1):
                        presentSheet.cell(row=sheetRowMark, column=j).font = Font(size=15, bold=True)
                    presentSheet.cell(row=sheetRowMark, column=j).alignment = alignment
                    presentSheet.cell(row=sheetRowMark, column=j).fill = pattern_fill
                sheetRowMark = sheetRowMark + 1

            font = Font(size=15, bold=True)
            alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)

            # 复制问题得分，空出加名字的行
            for k in range(quizNumber):  # row
                pattern_fill = PatternFill(fill_type="solid", fgColor=quiz_colors[0])
                if k > quizmarkstart[1]:
                    pattern_fill = PatternFill(fill_type="solid", fgColor="8cb5f9")
                if k > quizmarkstart[2]:
                    pattern_fill = PatternFill(fill_type="solid", fgColor="7ad694")
                if k > quizmarkstart[3]:
                    pattern_fill = PatternFill(fill_type="solid", fgColor="8ed7dd")
                if k > quizmarkstart[4]:
                    pattern_fill = PatternFill(fill_type="solid", fgColor="fdd868")
                for j in range(1, lengthcomment):  # col
                    presentSheet.cell(row=sheetRowMark, column=j).value = presentOriginal.cell(row=1 + k,
                                                                                               column=j).value
                    presentSheet.cell(row=sheetRowMark, column=j).font = font
                    presentSheet.cell(row=sheetRowMark, column=j).alignment = alignment
                    presentSheet.cell(row=sheetRowMark, column=j).fill = pattern_fill
                sheetRowMark = sheetRowMark + 1

            if LoopTime == 0:
                for m in range(1, 7):
                    presentSheet.cell(row=45, column=m).value = presentOriginal.cell(row=returnAvg,
                                                                                     column=m).value
                    presentSheet.merge_cells("A1:L1")
                    # 调整标题格样式
                    presentSheet.cell(row=1, column=1).value = 'Historical Comparison'
                    presentSheet.cell(row=1, column=1).font = Font(size=30, bold=True)
                    presentSheet.cell(row=1, column=1).alignment = Alignment(horizontal="center",
                                                                             vertical="center",
                                                                             wrap_text=True)
                    presentSheet.row_dimensions[1].height = 30
            # 复制自己的平均分到最终位置，准备绘制个人成长表格
            for m in range(1, 7):
                if m == 1:
                    presentSheet.cell(row=45 + LoopTime + 1, column=m).value = presentOriginal.cell(
                        row=returnAvg + 1, column=m).value + it[:it.find('_')]
                else:
                    presentSheet.cell(row=45 + LoopTime + 1, column=m).value = presentOriginal.cell(row=returnAvg + 1,
                                                                                                    column=m).value

            # 美化表格，调整表格属性字体等

            MainWorkBook.save(resultdir + saveTableName)
            # 一次完成后根据循环次数直接计算出报告结尾位置(只运行一次）
        sheetRowMarkLoop = sheetRowMark
        LoopTime = LoopTime + 1

    # 形成自己的维度对比图并存到表格开头

    for sheetNum in range(1, len(nameList) + 1):
        if ifweight=='y':
            chartDrawerSum('result\\'+saveTableName, sheetNum, nameList[sheetNum - 1], 44, len(processedWorkBookPaths),
                       len(processedWorkBookPaths), 1,nameList)
        else:
            chartDrawerSum('result\\'+saveTableName, sheetNum, nameList[sheetNum - 1], 44, len(processedWorkBookPaths),
                       len(processedWorkBookPaths), 0,nameList)
        # 加入图片
        picPath = cacheDir+'ChartHistory\\history_' + str(sheetNum) + '.jpg'
        img = Image(picPath)  # 选择你的图片
        MainWorkBook.worksheets[sheetNum].add_image(img, 'A2')
        MainWorkBook.save(resultdir + saveTableName)
        #     #把表格存到个人的分excel表中
        wb = openpyxl.load_workbook(resultdir + saveTableName, data_only=True)
        use_less = wb.sheetnames
        use_less.remove(nameList[sheetNum - 1])
        for i in use_less:
            wb.remove(wb[i])
        if ifweight=='y':
            wb.save(resultdir + '\\PersonalReport\\{}.xlsx'.format(nameList[sheetNum - 1]))
        else:
            wb.save(resultdir + '\\PersonalReport\\{}.xlsx'.format(nameList[sheetNum - 1]))

'''
封装整个处理过程
'''
def generateResult(workBookPaths,ifweight,nameList,weight):
    if ifweight == 'y':
        print('Start to generate result with weight.')
    else:
        print('Start to generate result without weight.')
    quizNumList = []
    QuizstartList = []
    avgList = []
    commentList = []

    for it in workBookPaths:

        # 跳过json文件
        if it.find('json') != -1:
            continue
        # 读取要处理的月份报表
        workbookit = openpyxl.load_workbook('RawData\\' + it, data_only=True)

        '''
        QuizNum:
        Quizstart:
        Quizend
        returnMainC,
        returnCommentM,
        returnAvgM,
        Ano
        '''
        # MRName=MRxxxx in filename.
        MRName = it[it.find('MR'):it.find('MR') + 6]

        # 把rawdata处理为cache中的分表信息方便后续整合
        if ifweight == 'y':
            QuizNum, Quizstart, Quizend, returnMainC, returnCommentM, returnAvgM, Ano = \
                MainChartProcess(workbookit, MRName, ifweight,nameList,weight)
        else:
            QuizNum, Quizstart, Quizend, returnMainC, returnCommentM, returnAvgM, Ano = \
                MainChartProcess(workbookit, MRName,ifweight,nameList,weight)

        # Error processing,find bad list item in table.
        if QuizNum == -1:
            print('Find Error,stop processing.')
            input("“Press Any Key to close terminal.”")
            sys.exit()

        quizNum = returnCommentM - 1

        # Mark the start position of this month's table message in sum table.
        quizNumList.append(quizNum)
        QuizstartList.append(Quizstart)
        avgList.append(returnAvgM)
        commentList.append(returnCommentM)

        print('Finished Pre Process {}'.format(MRName))

    print('Start to generate result without weight.')
    # 处理过后的表的quiz行索引是1-quiz num
    # 表comment行索引是returnCommentM+comment num
    # 数据是returnAvgM-returnAvgM+2
    # rawData是1 - memberNum+1
    # 生成无权结果

    # 把cache中的文件整合成总表
    if ifweight == 'y':
        sumTableGen(quizNumList, QuizstartList, avgList, commentList, Ano,ifweight,nameList)
        print('Finished generate result with weight.')
    else:
        sumTableGen(quizNumList, QuizstartList, avgList, commentList, Ano,ifweight,nameList)
        print('Finished generate result without weight.')